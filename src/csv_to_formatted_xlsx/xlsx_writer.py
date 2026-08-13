"""建立格式化 XLSX、保護文字欄位並安全發布輸出檔。

本模組接收已驗證的 CsvData，依 ColumnRule 寫入 Excel。最重要的不變條件是：
條碼等識別碼必須同時具有字串值、data_type="s" 與 number_format="@"。

輸出先寫入同目錄暫存檔，確認可由 openpyxl 重新開啟後才用 os.replace 發布；
任何失敗都不得留下看似完整的正式 XLSX。
"""

from __future__ import annotations

import os
import tempfile
import unicodedata
from datetime import time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .column_rules import build_rules, convert_value
from .models import ColumnKind, CsvData

# 樣式集中定義，未來調整配色或對齊時不需搜尋寫檔迴圈。
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
BODY_ALIGNMENT = Alignment(vertical="center")
TEXT_ALIGNMENT = Alignment(horizontal="left", vertical="center")


class XlsxWriteError(OSError):
    """Raised when a workbook cannot be safely published."""


def _set_text(cell: Cell, value: str) -> None:
    """強制儲存為 Excel 文字，並阻止等號開頭的內容成為公式。"""

    cell.value = value
    # 只設定 number_format 不夠；openpyxl 仍會把開頭 '=' 自動判定為公式。
    cell.data_type = "s"
    cell.number_format = "@"
    cell.alignment = TEXT_ALIGNMENT


def _display_width(value: object | None) -> int:
    """估算 Excel 欄寬；中日韓寬字元以約兩個拉丁字元計算。"""

    text = "" if value is None else str(value)
    return sum(2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1 for char in text)


def _safe_sheet_title(title: str) -> str:
    """移除 Excel 禁用字元並限制工作表名稱為 31 字。"""

    cleaned = "".join("_" if char in r"[]:*?/\\" else char for char in title).strip()
    return (cleaned or "報表")[:31]


def write_xlsx(
    data: CsvData,
    output_path: Path,
    *,
    overwrite: bool = True,
    protected_aliases: set[str] | None = None,
    sheet_title: str = "報表",
) -> list[str]:
    """寫入並驗證 XLSX，再以原子性取代發布至正式路徑。"""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists() and not overwrite:
        raise XlsxWriteError(f"輸出檔已存在：{output_path.name}")

    rules = build_rules(data.headers, protected_aliases)
    warnings = list(data.warnings)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_sheet_title(sheet_title)

    widths = [_display_width(header) for header in data.headers]
    for column, header in enumerate(data.headers, start=1):
        cell = worksheet.cell(row=1, column=column)
        _set_text(cell, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    # 限制逐格警告數量，避免品質很差的大檔案耗盡記憶體或塞爆日誌。
    warning_limit = 100
    for row_index, row in enumerate(data.rows, start=2):
        for column_index, (raw_value, rule) in enumerate(zip(row, rules, strict=True), start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            converted, warning = convert_value(raw_value, rule.kind)
            if warning and len(warnings) < warning_limit:
                warnings.append(f"第 {row_index} 列「{rule.header}」：{warning}")

            if converted is None:
                cell.value = None
                cell.alignment = BODY_ALIGNMENT
            elif isinstance(converted, str):
                _set_text(cell, converted)
            else:
                cell.value = converted
                cell.alignment = BODY_ALIGNMENT
                if rule.kind is ColumnKind.INTEGER:
                    cell.number_format = "#,##0"
                elif rule.kind is ColumnKind.DECIMAL:
                    cell.number_format = "#,##0.00"
                elif rule.kind is ColumnKind.PERCENT:
                    cell.number_format = "0.00%"
                elif rule.kind is ColumnKind.DATE:
                    cell.number_format = "yyyy-mm-dd"
                elif rule.kind is ColumnKind.DATETIME:
                    cell.number_format = (
                        "hh:mm:ss" if isinstance(converted, time) else "yyyy-mm-dd hh:mm:ss"
                    )
                elif rule.kind is ColumnKind.YEAR_MONTH:
                    cell.number_format = "yyyy-mm"

            # 只取前 500 筆資料估算欄寬，避免大型報表增加不必要的運算成本。
            if row_index <= 501:
                widths[column_index - 1] = max(
                    widths[column_index - 1], _display_width(raw_value)
                )

    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 24
    last_column = get_column_letter(len(data.headers))
    last_row = max(1, len(data.rows) + 1)
    worksheet.auto_filter.ref = f"A1:{last_column}{last_row}"
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(45, max(8, width + 2))

    # 暫存檔必須和正式檔在同一目錄，os.replace 才能可靠地原子性發布。
    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{output_path.stem}-", suffix=".tmp.xlsx", dir=output_path.parent, delete=False
        ) as temp_file:
            temp_path = Path(temp_file.name)
        workbook.save(temp_path)
        workbook.close()

        # 最低限度回讀可攔截損壞 ZIP/XLSX；data_only=False 也方便發現意外公式。
        verification = load_workbook(temp_path, read_only=True, data_only=False)
        try:
            if not verification.sheetnames:
                raise XlsxWriteError("暫存 XLSX 不含工作表")
        finally:
            verification.close()
        os.replace(temp_path, output_path)
    except Exception as exc:
        # 發布失敗時只刪暫存檔；既有正式輸出必須保持原狀。
        if temp_path is not None:
            try:
                temp_path.unlink(missing_ok=True)
            except OSError:
                pass
        if isinstance(exc, XlsxWriteError):
            raise
        raise XlsxWriteError(f"無法寫入 {output_path.name}：{exc}") from exc

    return warnings
