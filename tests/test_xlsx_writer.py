"""XLSX 寫入、格式、安全發布與回讀驗證測試。

這些測試直接重新開啟產出的工作簿，確認儲存格 data_type 與 number_format，
因為只檢查畫面顯示無法發現長條碼已失真或公式字串被 Excel 執行等問題。
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from csv_to_formatted_xlsx.models import CsvData
from csv_to_formatted_xlsx.xlsx_writer import XlsxWriteError, write_xlsx


def test_xlsx_preserves_identifiers_and_blocks_formulas(tmp_path: Path) -> None:
    """識別碼必須真正以文字儲存，公式外觀內容也不得成為公式。"""

    output = tmp_path / "report.xlsx"
    data = CsvData(
        headers=["條碼", "數量", "% Δ", "% Δ", "備註"],
        rows=[
            ["02131137", "1,234", "12.5%", "0.125", "=HYPERLINK(\"x\")"],
            ["123456789012345678", "無資料", "", "不明", "+SUM(1,2)"],
        ],
        encoding="utf-8",
    )

    warnings = write_xlsx(data, output)

    workbook = load_workbook(output, data_only=False)
    worksheet = workbook["報表"]
    # value、data_type、number_format 三者都要正確，才算完整保護條碼。
    assert worksheet["A2"].value == "02131137"
    assert worksheet["A2"].data_type == "s"
    assert worksheet["A2"].number_format == "@"
    assert worksheet["A3"].value == "123456789012345678"
    assert worksheet["A3"].data_type == "s"
    assert worksheet["B2"].value == 1234
    assert worksheet["B2"].number_format == "#,##0"
    assert worksheet["C2"].value == pytest.approx(0.125)
    assert worksheet["C2"].number_format == "0.00%"
    # data_type='s' 是公式注入防護的關鍵；只有顯示格式 '@' 並不足夠。
    assert worksheet["E2"].value == '=HYPERLINK("x")'
    assert worksheet["E2"].data_type == "s"
    assert worksheet["E3"].data_type == "s"
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:E3"
    workbook.close()
    assert any("數量" in warning for warning in warnings)
    assert any("百分比" in warning for warning in warnings)


def test_existing_output_is_untouched_when_overwrite_is_disabled(tmp_path: Path) -> None:
    """禁止覆蓋時，既有輸出內容即使不是有效 XLSX 也不得被改動。"""

    output = tmp_path / "report.xlsx"
    output.write_bytes(b"existing")
    data = CsvData(headers=["條碼"], rows=[["001"]], encoding="utf-8")

    with pytest.raises(XlsxWriteError, match="輸出檔已存在"):
        write_xlsx(data, output, overwrite=False)

    assert output.read_bytes() == b"existing"


def test_keyword_barcode_and_order_period_cells_are_pivot_friendly(tmp_path: Path) -> None:
    """條碼維持文字，同時讓年、月與日期時間可供樞紐辨識。"""

    output = tmp_path / "report.xlsx"
    data = CsvData(
        headers=["商品條碼（主要）", "Barcode Number", "訂單年", "月", "訂單時間", "訂單年月"],
        rows=[
            [
                "001234567890123456",
                "009988776655",
                "2026",
                "08",
                "2026-08-13 09:30",
                "2026-08",
            ]
        ],
        encoding="utf-8",
    )

    write_xlsx(data, output)

    workbook = load_workbook(output, data_only=False)
    worksheet = workbook.active
    # 條碼欄：s 代表字串，@ 代表 Excel 文字顯示格式。
    for coordinate in ("A2", "B2"):
        assert worksheet[coordinate].data_type == "s"
        assert worksheet[coordinate].number_format == "@"
    assert worksheet["A2"].value == "001234567890123456"
    # 年／月使用數值 n，避免樞紐以文字順序排列月份。
    assert worksheet["C2"].value == 2026
    assert worksheet["D2"].value == 8
    assert worksheet["C2"].data_type == "n"
    assert worksheet["D2"].data_type == "n"
    # 日期時間／年月使用日期 d；年月以當月一日儲存、畫面只顯示 yyyy-mm。
    assert worksheet["E2"].data_type == "d"
    assert worksheet["E2"].number_format == "yyyy-mm-dd hh:mm:ss"
    assert worksheet["F2"].data_type == "d"
    assert worksheet["F2"].number_format == "yyyy-mm"
    workbook.close()
